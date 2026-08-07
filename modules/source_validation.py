from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from modules.config import ConfigManager


@dataclass
class ColumnCheck:
    name: str
    ok: bool


@dataclass
class SourceValidation:
    name: str
    file_ok: bool
    sheet_ok: bool
    columns: list[ColumnCheck] = field(default_factory=list)
    file_path: str = ""
    sheet: str = ""
    message: str = ""
    available_sheets: list[str] = field(default_factory=list)
    available_columns: list[str] = field(default_factory=list)
    header_row_skip: int | None = None

    @property
    def columns_ok(self) -> bool:
        return bool(self.columns) and all(c.ok for c in self.columns)

    @property
    def ok(self) -> bool:
        return self.file_ok and self.sheet_ok and self.columns_ok


def _list_sheets(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _find_header(
    path: Path, sheet: str, required: list[str]
) -> tuple[int | None, list[str]]:
    """Return (skiprows, columns) when all required columns are present.

    If not all required columns match, return the header row that matches the
    most required names (usually skiprows=0) for clearer diagnostics.
    """
    best_skip: int | None = None
    best_cols: list[str] = []
    best_score = -1

    for skip in range(11):
        try:
            headers = pd.read_excel(path, sheet_name=sheet, skiprows=skip, nrows=0)
            cols = [str(c) for c in headers.columns.tolist()]
        except Exception:
            continue

        if not required:
            return skip, cols

        score = sum(1 for col in required if col in cols)
        if score > best_score:
            best_score = score
            best_skip = skip
            best_cols = cols

        if score == len(required):
            return skip, cols

    return (best_skip if best_score == len(required) else None), best_cols


def validate_source(
    name: str,
    entry: dict,
    project_root: str | Path | None = None,
) -> SourceValidation:
    required = [str(c).strip() for c in (entry.get("rows") or []) if str(c).strip()]
    sheet = (entry.get("sheet") or "").strip()
    resolved = ConfigManager.resolve_existing_path(entry.get("file_path"), project_root)

    if resolved is None:
        diag = ConfigManager.diagnose_path(entry.get("file_path"), project_root)
        return SourceValidation(
            name=name,
            file_ok=False,
            sheet_ok=False,
            columns=[ColumnCheck(c, False) for c in required],
            file_path=entry.get("file_path") or "",
            sheet=sheet,
            message=diag["message"],
        )

    file_path = str(resolved)
    try:
        sheets = _list_sheets(resolved)
    except Exception as exc:
        return SourceValidation(
            name=name,
            file_ok=True,
            sheet_ok=False,
            columns=[ColumnCheck(c, False) for c in required],
            file_path=file_path,
            sheet=sheet,
            message=f"Could not open workbook: {exc}",
        )

    if not sheet:
        return SourceValidation(
            name=name,
            file_ok=True,
            sheet_ok=False,
            columns=[ColumnCheck(c, False) for c in required],
            file_path=file_path,
            sheet=sheet,
            available_sheets=sheets,
            message="Sheet name is empty.",
        )

    if sheet not in sheets:
        return SourceValidation(
            name=name,
            file_ok=True,
            sheet_ok=False,
            columns=[ColumnCheck(c, False) for c in required],
            file_path=file_path,
            sheet=sheet,
            available_sheets=sheets,
            message=f"Sheet `{sheet}` not found. Available: {', '.join(sheets)}",
        )

    if not required:
        return SourceValidation(
            name=name,
            file_ok=True,
            sheet_ok=True,
            columns=[],
            file_path=file_path,
            sheet=sheet,
            available_sheets=sheets,
            message="No columns configured.",
        )

    skip, available = _find_header(resolved, sheet, required)
    column_checks = [ColumnCheck(col, col in available) for col in required]
    missing = [c.name for c in column_checks if not c.ok]

    if missing:
        return SourceValidation(
            name=name,
            file_ok=True,
            sheet_ok=True,
            columns=column_checks,
            file_path=file_path,
            sheet=sheet,
            available_sheets=sheets,
            available_columns=available,
            header_row_skip=skip,
            message=(
                f"Missing columns: {', '.join(missing)}. "
                f"Available on sheet: {', '.join(available) if available else '(none readable)'}"
            ),
        )

    return SourceValidation(
        name=name,
        file_ok=True,
        sheet_ok=True,
        columns=column_checks,
        file_path=file_path,
        sheet=sheet,
        available_sheets=sheets,
        available_columns=available,
        header_row_skip=skip,
        message="File, sheet, and columns OK.",
    )


def validate_section(
    section: dict | None,
    project_root: str | Path | None = None,
) -> list[SourceValidation]:
    section = section or {}
    return [validate_source(name, entry or {}, project_root) for name, entry in section.items()]


def section_ready(section: dict | None, project_root: str | Path | None = None) -> bool:
    validations = validate_section(section, project_root)
    return bool(validations) and all(v.ok for v in validations)
